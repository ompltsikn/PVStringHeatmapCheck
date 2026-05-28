"""Extend M2 PV Performance Workbook — Iterasi 3: M2bPeerZScore.

Input: docs/M2_PV_Performance_Workbook.xlsx (output Iterasi 2)
Output: same file, dengan 6 sheet baru:
  9.  Meteo_Dummy       — POA per timestamp (5 source columns: 1 dummy + placeholders)
  10. PanelSpec         — Jinko JKM625N datasheet + voc_at_cell_temp formula
  11. Raw_Data_M2b      — extended raw data dengan V, I per PV + sunrise/sunset timesteps
  12. Helpers_M2b       — per-row R_str, voc_actual estimation, POA mask
  13. M2b_PeerZScore    — per-(inv, PV) z_mean, z_median, voc_ratio, fault decision
  14. M2b_StringStatus  — replika Python output sheet
  15. Hampel_Preprocessing — A/B comparison V series original vs Hampel-cleaned
  16. M2b_StatComparison — mean vs median Z-score per PV (cross-check)

Sheet existing Iterasi 2 TIDAK diubah. Cell-level formula reproducible.
Skenario dummy diperluas: extend ke 24 timestep (06:00..07:55 sunrise window +
12:00..13:55 noon window) untuk demonstrasi voc_actual estimation di sunrise.
"""
from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule

INPUT = Path(__file__).parent / "M2_PV_Performance_Workbook.xlsx"

# --- Styling ---------------------------------------------------------------
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
SUBHEADER_FONT = Font(name="Calibri", size=10, bold=True)
SEV_FILL = {
    "CRITICAL": PatternFill("solid", fgColor="E06666"),
    "HIGH":     PatternFill("solid", fgColor="F6B26B"),
    "MEDIUM":   PatternFill("solid", fgColor="FFD966"),
    "INFO":     PatternFill("solid", fgColor="A4C2F4"),
    "NORMAL":   PatternFill("solid", fgColor="B6D7A8"),
    "EMPTY":    PatternFill("solid", fgColor="DDDDDD"),
    "high_R":   PatternFill("solid", fgColor="E06666"),
}


def set_header(ws, row, headers):
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


# Load existing workbook (Iterasi 2 output)
wb = load_workbook(INPUT)
existing_sheets = list(wb.sheetnames)
print(f"Existing sheets: {existing_sheets}")

# Add WB05-INV05 entry to EmptyPVMap (PV5 empty by design for M2b dummy)
ws_emp = wb["EmptyPVMap"]
# Find first empty row in EmptyPVMap (header at row 4, data starts row 5)
emp_row = 5
while ws_emp.cell(row=emp_row, column=1).value:
    emp_row += 1
ws_emp.cell(row=emp_row, column=1, value="WB05-INV05").border = BORDER
ws_emp.cell(row=emp_row, column=2, value="5").border = BORDER
ws_emp.cell(row=emp_row, column=3, value="DEMO M2b: PV5 EMPTY").border = BORDER
# Compute new range end (was 8 before, now emp_row)
EMPTY_MAP_END = emp_row

# ===========================================================================
# Update README iterasi log
# ===========================================================================
ws = wb["README"]
# Find last log row
log_row = None
for r in range(5, 20):
    if ws.cell(row=r, column=1).value and "Iterasi" in str(ws.cell(row=r, column=1).value or ""):
        continue
    if not ws.cell(row=r, column=1).value:
        log_row = r
        break
if log_row:
    ws.cell(row=log_row, column=1, value="3").border = BORDER
    ws.cell(row=log_row, column=2, value="2026-05-28").border = BORDER
    ws.cell(row=log_row, column=3, value="M2bPeerZScore").border = BORDER
    ws.cell(row=log_row, column=4, value="Meteo_Dummy, PanelSpec, Raw_Data_M2b, Helpers_M2b, M2b_PeerZScore, M2b_StringStatus, Hampel_Preprocessing, M2b_StatComparison").border = BORDER


# ===========================================================================
# Extend Config sheet — add M2b thresholds + defined names
# ===========================================================================
ws = wb["Config"]
# Find last config row
last_cfg_row = 16  # known from Iterasi 2 (12 rows + 3-row offset)
# Actually iterate to find truly the last row with section name
r = 4
while ws.cell(row=r, column=1).value:
    last_cfg_row = r
    r += 1
start_row = last_cfg_row + 1

new_cfg = [
    # m2b core
    ("m2b", "poa_threshold_wm2",        300.0, "W/m²; |z| Z-score hanya hitung saat POA > ini (spec 4.2.1)"),
    ("m2b", "poa_floor_wm2",            50.0,  "W/m²; sunset fix — POA harus > ini (sensor lag protection)"),
    ("m2b", "z_threshold",              2.5,   "|z| > ini → flag suspect (spec 4.2.1)"),
    ("m2b", "z_high_threshold",         3.5,   "|z| > ini → severity HIGH; else MEDIUM"),
    ("m2b", "voc_ratio_threshold",      0.95,  "voc_actual/voc_string_nominal > ini → confirm high_R (spec 4.2.3)"),
    ("m2b", "stat_method",              "median", "mean | median | both (z_primary picker)"),
    ("m2b", "min_peer_strings",         3,     "min sibling per inverter untuk valid peer comparison"),
    ("m2b", "min_daylight_samples",     10,    "min samples lulus POA gate per inverter"),
    ("m2b", "i_clip_floor_a",           0.1,   "clip(lower=0.1A) untuk hindari divide-by-zero saat R=V/I"),
    # voc estimator
    ("voc_estimator", "i_threshold_a",  0.5,   "I < ini → dianggap open-circuit (sunrise/sunset)"),
    ("voc_estimator", "min_voc_v",      10.0,  "V > ini → buang zero-reading sample"),
    ("voc_estimator", "min_samples",    3,     "minimum samples valid untuk median voc"),
    # Hampel preprocessing
    ("preprocessing", "hampel_window",        15,  "samples; rolling window (75 menit @ 5-min)"),
    ("preprocessing", "hampel_max_deviation", 3.0, "MAD-sigma; |x - median| / MAD > ini → outlier"),
    ("preprocessing", "enabled",              0,   "1=apply Hampel cleaning, 0=skip (raw V/I)"),
    # POA single-source (dummy)
    ("poa", "default_source",           "pyranometer_avg", "single source di dummy (production: 5 sources)"),
]

for i, (section, key, value, note) in enumerate(new_cfg, start=start_row):
    ws.cell(row=i, column=1, value=section).border = BORDER
    ws.cell(row=i, column=2, value=key).border = BORDER
    c = ws.cell(row=i, column=3, value=value)
    c.border = BORDER
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="center")
    ws.cell(row=i, column=4, value=note).border = BORDER
    name = f"cfg_{key}"
    if name not in wb.defined_names:
        dn = DefinedName(name, attr_text=f"Config!$C${i}")
        wb.defined_names[name] = dn


# ===========================================================================
# Sheet: PanelSpec — Jinko JKM625N datasheet
# ===========================================================================
ws = wb.create_sheet("PanelSpec")
ws["A1"] = "PanelSpec — Jinko JKM625N 78HL4-BDV (mirror config/panel_spec.yaml)"
ws["A1"].font = Font(size=14, bold=True, color="305496")
ws["A2"] = "Konsumen: PanelSpec.voc_at_cell_temp + voc_string_nominal di peer_zscore.py"
ws["A2"].font = Font(italic=True, color="595959")

set_header(ws, 4, ["Parameter", "Value", "Unit", "Notes"])
panel_rows = [
    ("panel_model", "Jinko JKM625N 78HL4-BDV", "", "Bifacial Tiger Neo dual-glass"),
    ("voc_stc_v", 55.72, "V", "Open-circuit voltage @ STC (1000 W/m², 25°C cell)"),
    ("vmp_stc_v", 46.10, "V", "Max-power voltage @ STC"),
    ("imp_stc_a", 13.56, "A", "Max-power current @ STC"),
    ("isc_stc_a", 14.27, "A", "Short-circuit current @ STC"),
    ("pmax_stc_w", 625, "W", "Max power @ STC"),
    ("voc_noct_v", 52.93, "V", "Voc @ NOCT (800 W/m², 20°C ambient)"),
    ("temp_coef_voc_pct_per_c", -0.25, "%/°C", "Voc decreases when Tcell rises"),
    ("temp_coef_pmax_pct_per_c", -0.29, "%/°C", "Pmax temperature coefficient"),
    ("temp_coef_isc_pct_per_c", 0.045, "%/°C", "Isc temperature coefficient (small +)"),
    ("t_stc_c", 25.0, "°C", "STC reference temperature"),
    ("max_system_voltage_v", 1500, "V", "IEC limit"),
    ("default_modules_per_string", 26, "modules", "Fallback (WB03-10)"),
    ("modules_per_string_WB01", 24, "modules", "PLTS-IKN WB01 layout"),
    ("modules_per_string_WB02", 24, "modules", "PLTS-IKN WB02 layout"),
    ("modules_per_string_WB03_10", 26, "modules", "PLTS-IKN WB03-10 layout"),
]
named_panel = {}
for i, (param, val, unit, note) in enumerate(panel_rows, start=5):
    ws.cell(row=i, column=1, value=param).border = BORDER
    c = ws.cell(row=i, column=2, value=val)
    c.border = BORDER
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="center")
    ws.cell(row=i, column=3, value=unit).border = BORDER
    ws.cell(row=i, column=4, value=note).border = BORDER
    if isinstance(val, (int, float)):
        named_panel[f"panel_{param}"] = f"PanelSpec!$B${i}"

# Define panel named ranges
for name, ref in named_panel.items():
    if name not in wb.defined_names:
        dn = DefinedName(name, attr_text=ref)
        wb.defined_names[name] = dn

# Voc formula calculator (helper)
ws["A22"] = "voc_at_cell_temp(T_cell) — Excel formula"
ws["A22"].font = SUBHEADER_FONT
ws["A22"].fill = SUBHEADER_FILL
ws["A23"] = "Input: cell_temp_c →"
ws["B23"] = 30.0  # default Tcell 30°C
ws["B23"].font = Font(bold=True)
ws["B23"].fill = PatternFill("solid", fgColor="FFF2CC")
ws["C23"] = "°C (edit untuk recalc)"
ws["A24"] = "voc_per_module = voc_stc × (1 + temp_coef_voc/100 × (T - 25))"
ws["B24"] = "=panel_voc_stc_v*(1+panel_temp_coef_voc_pct_per_c/100*(B23-panel_t_stc_c))"
ws["B24"].number_format = "0.00"
ws["C24"] = "V per modul"
ws["A25"] = "voc_string_WB01_WB02 (24 modules)"
ws["B25"] = "=B24*panel_modules_per_string_WB01"
ws["B25"].number_format = "0.00"
ws["C25"] = "V per string (cold Voc skenario)"
ws["A26"] = "voc_string_WB03_WB10 (26 modules)"
ws["B26"] = "=B24*panel_default_modules_per_string"
ws["B26"].number_format = "0.00"
ws["C26"] = "V per string"

named_panel["voc_per_module_calc"] = "PanelSpec!$B$24"
named_panel["voc_string_24_calc"] = "PanelSpec!$B$25"
named_panel["voc_string_26_calc"] = "PanelSpec!$B$26"
for name, ref in [("voc_per_module_calc", "PanelSpec!$B$24"),
                   ("voc_string_24_calc", "PanelSpec!$B$25"),
                   ("voc_string_26_calc", "PanelSpec!$B$26"),
                   ("tcell_dummy_c", "PanelSpec!$B$23")]:
    if name not in wb.defined_names:
        dn = DefinedName(name, attr_text=ref)
        wb.defined_names[name] = dn

for letter, w in zip("ABCD", [40, 15, 12, 60]):
    ws.column_dimensions[letter].width = w


# ===========================================================================
# Sheet: Raw_Data_M2b — extended dummy data (24 timestep)
# ===========================================================================
# Skenario:
#   - 1 inverter WB05-INV05 (26 modules/string per panel_spec)
#   - 5 PV strings (PV1..PV5)
#   - 24 timestep: 06:00..06:55 (sunrise window, I≈0) + 12:00..13:55 (noon, full sun)
#   - PV3: HIGH-R fault — V slightly elevated, I significantly reduced (60% reduction).
#   - PV5: EMPTY (no V/I data — placeholder zero)
#   - PV1, PV2, PV4: normal
#
# Realistic Jinko 26-module string:
#   STC: Voc=55.72×26=1448.7V, Imp≈13A, Vmp≈1200V
#   Sunrise: V≈1400 (cold Voc), I≈0
#   Noon: V≈1200 (Vmp), I≈13A

ws = wb.create_sheet("Raw_Data_M2b")
ws["A1"] = "Raw_Data_M2b — extended dummy data (WB05-INV05, 24 timestep)"
ws["A1"].font = Font(size=14, bold=True, color="305496")
ws["A2"] = ("Skenario: 06:00-06:55 sunrise (I≈0, V≈Voc), 12:00-13:55 noon (full sun). "
            "PV3 punya HIGH-R fault (I drop 60%, V naik sedikit) → expected flag.")
ws["A2"].font = Font(italic=True, color="595959")

headers_m2b = ["Inverter_ID", "Start Time", "Inverter status"]
for pv_n in range(1, 6):
    headers_m2b += [f"PV{pv_n} input voltage(V)", f"PV{pv_n} input current(A)"]
set_header(ws, 4, headers_m2b)

import math
# Generate 24 timesteps: 06:00..06:55 (12 ts) + 12:00..13:55 (24 ts → take first 12 noon)
# Actually use 24 total: 06:00, 06:05, ..., 06:55 (12 sunrise) + 12:00..12:55 (12 noon)
sunrise_start = datetime(2026, 5, 14, 6, 0, 0)
noon_start    = datetime(2026, 5, 14, 12, 0, 0)
times = [sunrise_start + timedelta(minutes=5*t) for t in range(12)] + \
        [noon_start    + timedelta(minutes=5*t) for t in range(12)]

# Voc cold @ 20°C ≈ 55.72 × (1 + (-0.25/100) × (20-25)) × 26 = 55.72 × 1.0125 × 26 = 1467.3V
# Vmp noon @ 50°C ≈ 1200V
# Imp noon = 13A
import random
random.seed(42)
def jitter(scale): return random.uniform(-scale, scale)

raw_rows = []
for ts in times:
    is_sunrise = ts.hour < 12
    if is_sunrise:
        # Sunrise: V near Voc cold (open circuit), I tiny
        v_base = 1467 + jitter(5)
        i_base = 0.1 + jitter(0.1)
    else:
        # Noon: full sun
        v_base = 1200 + jitter(8)
        i_base = 13.0 + jitter(0.15)
    pv_vals = {}
    for pv_n in range(1, 6):
        if pv_n == 5:
            # EMPTY
            pv_vals[pv_n] = (0.0, 0.0)
        elif pv_n == 3 and not is_sunrise:
            # HIGH-R fault: V slightly elevated (less current draw → higher V), I much lower
            pv_vals[pv_n] = (v_base * 1.03 + jitter(2), i_base * 0.40 + jitter(0.05))
        else:
            pv_vals[pv_n] = (v_base + jitter(2), i_base + jitter(0.05))
    row = ["WB05-INV05", ts, "Grid connected"]
    for pv_n in range(1, 6):
        v, i = pv_vals[pv_n]
        row.append(round(v, 2))
        row.append(round(i, 3))
    raw_rows.append(row)

for ri, row_tuple in enumerate(raw_rows, start=5):
    for ci, v in enumerate(row_tuple, start=1):
        c = ws.cell(row=ri, column=ci, value=v)
        c.border = BORDER
        if ci == 2:
            c.number_format = "yyyy-mm-dd hh:mm:ss"

ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 18
for col_idx in range(4, len(headers_m2b)+1):
    ws.column_dimensions[get_column_letter(col_idx)].width = 13

LAST_M2B = 4 + len(raw_rows)  # = 28 (4 header + 24 data)


# ===========================================================================
# Sheet: Meteo_Dummy — POA per timestamp (single source for simplicity)
# ===========================================================================
ws = wb.create_sheet("Meteo_Dummy")
ws["A1"] = "Meteo_Dummy — POA (Plane-of-Array) per timestamp"
ws["A1"].font = Font(size=14, bold=True, color="305496")
ws["A2"] = ("Single source 'pyranometer_avg' (dummy). Produksi: 5 source fan-out "
            "(pyranometer_per_ws, pyranometer_avg, pvlib_clearsky_ineichen/simplified_solis/haurwitz).")
ws["A2"].font = Font(italic=True, color="595959")

set_header(ws, 4, ["Start Time", "POA pyranometer_avg (W/m²)", "Notes"])
# POA values:
#   06:00..06:55: 30..80 W/m² (sunrise ramp, NOT meets POA threshold 300)
#   12:00..13:55: 800..950 W/m² (full sun, meets gate)
poa_vals = []
for i, ts in enumerate(times):
    if ts.hour < 12:
        # ramp from 30 to 80 over 12 timesteps
        poa = 30 + i * (50/11)
    else:
        # 800..900 with slight variation
        poa = 850 + (i - 12) * 5 + jitter(15)
    poa_vals.append(poa)

for ri, (ts, poa) in enumerate(zip(times, poa_vals), start=5):
    ws.cell(row=ri, column=1, value=ts).border = BORDER
    ws.cell(row=ri, column=1).number_format = "yyyy-mm-dd hh:mm:ss"
    c = ws.cell(row=ri, column=2, value=round(poa, 1))
    c.border = BORDER
    c.number_format = "0.0"
    note = "sunrise (below gate)" if ts.hour < 12 else "noon (above gate)"
    ws.cell(row=ri, column=3, value=note).border = BORDER

ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 24


# ===========================================================================
# Sheet: Helpers_M2b — per-row R_str + masks + Hampel preprocessing
# ===========================================================================
ws = wb.create_sheet("Helpers_M2b")
ws["A1"] = "Helpers_M2b — derived columns per row dari Raw_Data_M2b"
ws["A1"].font = Font(size=14, bold=True, color="305496")
ws["A2"] = ("Per-row: POA value (VLOOKUP dari Meteo_Dummy), mask_poa, R_str per PV. "
            "Aggregate (R_str median per PV, fleet stats) ada di sheet M2b_PeerZScore.")
ws["A2"].font = Font(italic=True, color="595959")

helper_headers = [
    "Inverter_ID", "Start Time", "status",          # 1-3 mirror raw
    "POA_Wm2",                                       # 4 VLOOKUP from Meteo_Dummy
    "mask_poa",                                      # 5 POA > threshold AND > floor
    "is_on",                                         # 6
    # Per-PV: V, I, R_str=V/I_clip, voc_candidate(I<thr & V>min_voc)
    "PV1_V", "PV1_I", "PV1_R", "PV1_voc_cand",     # 7-10
    "PV2_V", "PV2_I", "PV2_R", "PV2_voc_cand",     # 11-14
    "PV3_V", "PV3_I", "PV3_R", "PV3_voc_cand",     # 15-18
    "PV4_V", "PV4_I", "PV4_R", "PV4_voc_cand",     # 19-22
    "PV5_V", "PV5_I", "PV5_R", "PV5_voc_cand",     # 23-26
]
set_header(ws, 4, helper_headers)
ws.freeze_panes = "D5"

# Per-row formulas. Raw_Data_M2b rows: 5..28.
for ri in range(5, LAST_M2B + 1):
    raw = ri  # same row
    # 1-3 mirror
    ws.cell(row=ri, column=1, value=f"=Raw_Data_M2b!A{raw}").border = BORDER
    c = ws.cell(row=ri, column=2, value=f"=Raw_Data_M2b!B{raw}")
    c.border = BORDER
    c.number_format = "yyyy-mm-dd hh:mm:ss"
    ws.cell(row=ri, column=3, value=f"=Raw_Data_M2b!C{raw}").border = BORDER

    # 4: POA via VLOOKUP from Meteo_Dummy
    poa_formula = f"=IFERROR(VLOOKUP(B{ri},Meteo_Dummy!$A$5:$B${4+len(times)},2,FALSE),0)"
    c = ws.cell(row=ri, column=4, value=poa_formula)
    c.border = BORDER
    c.number_format = "0.0"

    # 5: mask_poa = (POA > poa_threshold AND POA > poa_floor)
    ws.cell(row=ri, column=5,
            value=f"=IF(AND(D{ri}>cfg_poa_threshold_wm2,D{ri}>cfg_poa_floor_wm2),1,0)").border = BORDER

    # 6: is_on
    ws.cell(row=ri, column=6,
            value=f'=IF(ISNUMBER(SEARCH("grid connected",LOWER(C{ri}))),1,0)').border = BORDER

    # Per-PV: V (cols 7,11,15,19,23), I (8,12,16,20,24), R (9,13,17,21,25), voc_cand (10,14,18,22,26)
    pv_letters = {
        1: ("D", "E"),   # Raw_Data_M2b PV1: col D (V), col E (I)
        2: ("F", "G"),
        3: ("H", "I"),
        4: ("J", "K"),
        5: ("L", "M"),
    }
    for pv_n, (v_letter, i_letter) in pv_letters.items():
        col_v = 3 + (pv_n - 1) * 4 + 4   # 7, 11, 15, 19, 23
        col_i = col_v + 1
        col_r = col_v + 2
        col_voc = col_v + 3
        # V
        ws.cell(row=ri, column=col_v,
                value=f"=Raw_Data_M2b!{v_letter}{raw}").border = BORDER
        ws.cell(row=ri, column=col_v).number_format = "0.00"
        # I
        ws.cell(row=ri, column=col_i,
                value=f"=Raw_Data_M2b!{i_letter}{raw}").border = BORDER
        ws.cell(row=ri, column=col_i).number_format = "0.000"
        # R_str = V / max(I, i_clip_floor), only valid when mask_poa=1
        v_addr = f"{get_column_letter(col_v)}{ri}"
        i_addr = f"{get_column_letter(col_i)}{ri}"
        r_formula = f"=IF(E{ri}=1, {v_addr}/MAX({i_addr},cfg_i_clip_floor_a), \"\")"
        ws.cell(row=ri, column=col_r, value=r_formula).border = BORDER
        ws.cell(row=ri, column=col_r).number_format = "0.0"
        # voc_cand = (I < i_threshold_voc AND V > min_voc) ? V : ""
        voc_formula = (
            f"=IF(AND(ABS({i_addr})<cfg_i_threshold_a, {v_addr}>cfg_min_voc_v), {v_addr}, \"\")"
        )
        ws.cell(row=ri, column=col_voc, value=voc_formula).border = BORDER
        ws.cell(row=ri, column=col_voc).number_format = "0.00"

ws.column_dimensions["A"].width = 13
ws.column_dimensions["B"].width = 19
ws.column_dimensions["C"].width = 17
for col_idx in range(4, len(helper_headers)+1):
    ws.column_dimensions[get_column_letter(col_idx)].width = 10


# ===========================================================================
# Sheet: M2b_PeerZScore — per-PV aggregate + fault decision
# ===========================================================================
ws = wb.create_sheet("M2b_PeerZScore")
ws["A1"] = "M2b PeerZScore — agregat per PV (R_str median dari mask_poa samples) + Z-score + voc_ratio + fault decision"
ws["A1"].font = Font(size=14, bold=True, color="305496")
ws["A2"] = ("Spec 4.2.1: |z| > 2.5 → flag. Spec 4.2.3 high_R: AND voc_ratio > 0.95 → emit. "
            "Confidence = min(90, |z|/4*100). Severity: |z| > 3.5 HIGH else MEDIUM.")
ws["A2"].font = Font(italic=True, color="595959")

# --- Section A: per-PV metrics ---
ws["A4"] = "A. Per-PV string metrics (WB05-INV05, 5 PV strings)"
ws["A4"].font = SUBHEADER_FONT
ws["A4"].fill = SUBHEADER_FILL

set_header(ws, 5, [
    "PV_string", "R_str (median)", "voc_actual_v (median I<0.5A)",
    "voc_string_nominal_v", "voc_ratio",
    "n_R_samples", "n_voc_samples",
])

# Pull R_str median per PV (only mask_poa=1 samples) + voc_actual (only voc_cand non-empty)
pv_data_start = 6
for i, pv_n in enumerate(range(1, 6)):
    r = pv_data_start + i
    # Each PV occupies col_r and col_voc in Helpers_M2b
    col_r = 3 + (pv_n - 1) * 4 + 4 + 2   # 9, 13, 17, 21, 25
    col_voc = col_r + 1                   # 10, 14, 18, 22, 26
    helper_r_range = f"Helpers_M2b!${get_column_letter(col_r)}$5:${get_column_letter(col_r)}${LAST_M2B}"
    helper_voc_range = f"Helpers_M2b!${get_column_letter(col_voc)}$5:${get_column_letter(col_voc)}${LAST_M2B}"

    ws.cell(row=r, column=1, value=f"PV{pv_n}").border = BORDER

    # EMPTY check via EmptyPVMap lookup (mirror M2e behavior).
    # Built via .format() to avoid nested-quote f-string parser issues.
    is_empty_formula = (
        'IFERROR(IF(ISNUMBER(SEARCH('
        '","&{pv}&",",'
        '","&VLOOKUP("WB05-INV05",EmptyPVMap!$A$5:$B${end},2,FALSE)&","'
        ")),1,0),0)"
    ).format(pv=pv_n, end=EMPTY_MAP_END)

    # R_str: if EMPTY return text "EMPTY", else MEDIAN of helper range
    rstr_formula = (
        '=IF(' + is_empty_formula + '=1,"EMPTY",'
        'IFERROR(MEDIAN(' + helper_r_range + '),""))'
    )
    ws.cell(row=r, column=2, value=rstr_formula).border = BORDER
    ws.cell(row=r, column=2).number_format = "0.0"

    # voc_actual: if EMPTY return "EMPTY", else median of voc_cand
    voc_formula = (
        '=IF(' + is_empty_formula + '=1,"EMPTY",'
        'IFERROR(MEDIAN(' + helper_voc_range + '),""))'
    )
    ws.cell(row=r, column=3, value=voc_formula).border = BORDER
    ws.cell(row=r, column=3).number_format = "0.00"

    # voc_string_nominal — pakai PanelSpec voc_string_26_calc untuk WB05 (26 modules)
    ws.cell(row=r, column=4, value=f"=voc_string_26_calc").border = BORDER
    ws.cell(row=r, column=4).number_format = "0.00"

    # voc_ratio = voc_actual / voc_string_nominal
    ws.cell(row=r, column=5,
            value=f'=IFERROR(C{r}/D{r},"")').border = BORDER
    ws.cell(row=r, column=5).number_format = "0.000"

    # n_R_samples
    ws.cell(row=r, column=6,
            value=f"=COUNT({helper_r_range})").border = BORDER

    # n_voc_samples
    ws.cell(row=r, column=7,
            value=f"=COUNT({helper_voc_range})").border = BORDER


# --- Section B: Fleet stats (mean/median/std across all PVs) ---
fleet_row = pv_data_start + 5 + 2
ws.cell(row=fleet_row, column=1, value="B. Fleet statistics across PV1..PV5 (skip blanks)").font = SUBHEADER_FONT
ws.cell(row=fleet_row, column=1).fill = SUBHEADER_FILL

set_header(ws, fleet_row + 1, ["Metric", "Value", "", "", "", "", ""])
ws.cell(row=fleet_row+2, column=1, value="R_str fleet mean").border = BORDER
ws.cell(row=fleet_row+2, column=2,
        value=f"=AVERAGE(B{pv_data_start}:B{pv_data_start+4})").border = BORDER
ws.cell(row=fleet_row+2, column=2).number_format = "0.00"

ws.cell(row=fleet_row+3, column=1, value="R_str fleet median").border = BORDER
ws.cell(row=fleet_row+3, column=2,
        value=f"=MEDIAN(B{pv_data_start}:B{pv_data_start+4})").border = BORDER
ws.cell(row=fleet_row+3, column=2).number_format = "0.00"

ws.cell(row=fleet_row+4, column=1, value="R_str fleet std (sample stdev)").border = BORDER
# STDEV is universal-compatible (Excel 2010+, LibreOffice); STDEV.S returns #NAME? in older LibreOffice
ws.cell(row=fleet_row+4, column=2,
        value=f"=STDEV(B{pv_data_start}:B{pv_data_start+4})").border = BORDER
ws.cell(row=fleet_row+4, column=2).number_format = "0.00"

# Define named cells for fleet stats
for name, ref in [("rstr_fleet_mean", f"M2b_PeerZScore!$B${fleet_row+2}"),
                   ("rstr_fleet_median", f"M2b_PeerZScore!$B${fleet_row+3}"),
                   ("rstr_fleet_std", f"M2b_PeerZScore!$B${fleet_row+4}")]:
    if name not in wb.defined_names:
        dn = DefinedName(name, attr_text=ref)
        wb.defined_names[name] = dn


# --- Section C: Z-score & fault decision per PV ---
decision_row = fleet_row + 7
ws.cell(row=decision_row, column=1, value="C. Z-score & fault decision per PV").font = SUBHEADER_FONT
ws.cell(row=decision_row, column=1).fill = SUBHEADER_FILL

set_header(ws, decision_row + 1, [
    "PV_string", "z_mean", "z_median", "z_primary (per stat_method)",
    "flagged_by_z (|z|>2.5)", "voc_ok (voc_ratio>0.95)",
    "emit_finding", "fault_type", "severity", "confidence_pct", "message",
])

for i, pv_n in enumerate(range(1, 6)):
    r = decision_row + 2 + i
    pv_metrics_row = pv_data_start + i  # row di section A
    ws.cell(row=r, column=1, value=f"PV{pv_n}").border = BORDER

    # z_mean = (R_str - fleet_mean) / fleet_std
    ws.cell(row=r, column=2,
            value=f"=IFERROR((B{pv_metrics_row}-rstr_fleet_mean)/rstr_fleet_std,\"\")").border = BORDER
    ws.cell(row=r, column=2).number_format = "0.00"

    # z_median = (R_str - fleet_median) / fleet_std
    ws.cell(row=r, column=3,
            value=f"=IFERROR((B{pv_metrics_row}-rstr_fleet_median)/rstr_fleet_std,\"\")").border = BORDER
    ws.cell(row=r, column=3).number_format = "0.00"

    # z_primary based on stat_method
    ws.cell(row=r, column=4, value=(
        f'=IF(cfg_stat_method="median",C{r},'
        f'IF(cfg_stat_method="mean",B{r},'
        f'IF(ABS(B{r})>=ABS(C{r}),B{r},C{r})))'
    )).border = BORDER
    ws.cell(row=r, column=4).number_format = "0.00"

    # flagged_by_z
    ws.cell(row=r, column=5,
            value=f"=IF(ABS(D{r})>cfg_z_threshold,1,0)").border = BORDER

    # voc_ok = voc_ratio > 0.95
    voc_ratio_ref = f"E{pv_metrics_row}"  # col E in section A is voc_ratio
    ws.cell(row=r, column=6,
            value=f'=IFERROR(IF({voc_ratio_ref}>cfg_voc_ratio_threshold,1,0),0)').border = BORDER

    # emit_finding = flagged AND voc_ok
    ws.cell(row=r, column=7,
            value=f"=IF(AND(E{r}=1,F{r}=1),1,0)").border = BORDER

    # fault_type
    ws.cell(row=r, column=8,
            value=f'=IF(G{r}=1,"high_R","")').border = BORDER

    # severity
    ws.cell(row=r, column=9, value=(
        f'=IF(G{r}=1,IF(ABS(D{r})>cfg_z_high_threshold,"HIGH","MEDIUM"),"")'
    )).border = BORDER

    # confidence_pct = min(90, |z|/4*100)
    ws.cell(row=r, column=10,
            value=f'=IF(G{r}=1,MIN(90,ABS(D{r})/4*100),"")').border = BORDER
    ws.cell(row=r, column=10).number_format = "0.0"

    # message
    ws.cell(row=r, column=11, value=(
        f'=IF(G{r}=1,'
        f'"High-R suspect PV"&{pv_n}&": |z|="&TEXT(ABS(D{r}),"0.00")&" voc_ratio="&TEXT({voc_ratio_ref},"0.000"),'
        f'"")'
    )).border = BORDER

# Conditional formatting severity
sev_range_c = f"I{decision_row+2}:I{decision_row+6}"
for sev_name, fill in SEV_FILL.items():
    ws.conditional_formatting.add(
        sev_range_c,
        CellIsRule(operator="equal", formula=[f'"{sev_name}"'], fill=fill),
    )

for letter, w in zip("ABCDEFGHIJK", [12, 14, 25, 16, 14, 17, 17, 14, 11, 14, 70]):
    ws.column_dimensions[letter].width = w


# ===========================================================================
# Sheet: M2b_StringStatus — replika Python output
# ===========================================================================
ws = wb.create_sheet("M2b_StringStatus")
ws["A1"] = "M2b_StringStatus — replika Python self.artifacts['StringStatus']"
ws["A1"].font = Font(size=14, bold=True, color="305496")
ws["A2"] = "Pull dari M2b_PeerZScore sections A & C."
ws["A2"].font = Font(italic=True, color="595959")

set_header(ws, 4, [
    "poa_source", "inverter_id", "wb_id", "pv_string", "status",
    "rstr", "rstr_fleet_median", "rstr_fleet_std",
    "z_mean", "z_median", "voc_actual_v", "voc_string_nominal_v",
    "voc_ratio", "voc_ok", "emitted_finding",
])

for i in range(5):  # 5 PV
    r = 5 + i
    pv_metrics_row = pv_data_start + i
    decision_data_row = decision_row + 2 + i
    pv_n = i + 1

    ws.cell(row=r, column=1, value="pyranometer_avg").border = BORDER
    ws.cell(row=r, column=2, value="WB05-INV05").border = BORDER
    ws.cell(row=r, column=3, value="WB05").border = BORDER
    ws.cell(row=r, column=4, value=f"PV{pv_n}").border = BORDER
    # status: high_R if emit, EMPTY if PV5 (no data), else NORMAL
    if pv_n == 5:
        ws.cell(row=r, column=5, value="EMPTY").border = BORDER
    else:
        ws.cell(row=r, column=5,
                value=f'=IF(M2b_PeerZScore!G{decision_data_row}=1,"high_R","NORMAL")').border = BORDER
    ws.cell(row=r, column=6, value=f"=M2b_PeerZScore!B{pv_metrics_row}").border = BORDER
    ws.cell(row=r, column=6).number_format = "0.0"
    ws.cell(row=r, column=7, value="=rstr_fleet_median").border = BORDER
    ws.cell(row=r, column=7).number_format = "0.0"
    ws.cell(row=r, column=8, value="=rstr_fleet_std").border = BORDER
    ws.cell(row=r, column=8).number_format = "0.0"
    ws.cell(row=r, column=9, value=f"=M2b_PeerZScore!B{decision_data_row}").border = BORDER
    ws.cell(row=r, column=9).number_format = "0.00"
    ws.cell(row=r, column=10, value=f"=M2b_PeerZScore!C{decision_data_row}").border = BORDER
    ws.cell(row=r, column=10).number_format = "0.00"
    ws.cell(row=r, column=11, value=f"=M2b_PeerZScore!C{pv_metrics_row}").border = BORDER
    ws.cell(row=r, column=11).number_format = "0.00"
    ws.cell(row=r, column=12, value=f"=M2b_PeerZScore!D{pv_metrics_row}").border = BORDER
    ws.cell(row=r, column=12).number_format = "0.00"
    ws.cell(row=r, column=13, value=f"=M2b_PeerZScore!E{pv_metrics_row}").border = BORDER
    ws.cell(row=r, column=13).number_format = "0.000"
    ws.cell(row=r, column=14, value=f"=M2b_PeerZScore!F{decision_data_row}").border = BORDER
    ws.cell(row=r, column=15, value=f"=M2b_PeerZScore!G{decision_data_row}").border = BORDER

sev_range_st = f"E5:E9"
for sev_name, fill in [("high_R", SEV_FILL["high_R"]), ("NORMAL", SEV_FILL["NORMAL"]),
                       ("EMPTY", SEV_FILL["EMPTY"])]:
    ws.conditional_formatting.add(
        sev_range_st,
        CellIsRule(operator="equal", formula=[f'"{sev_name}"'], fill=fill),
    )

for letter, w in zip("ABCDEFGHIJKLMNO", [17, 14, 8, 11, 11, 9, 14, 12, 9, 10, 13, 17, 11, 9, 14]):
    ws.column_dimensions[letter].width = w


# ===========================================================================
# Sheet: M2b_StatComparison — median vs mean cross-check
# ===========================================================================
ws = wb.create_sheet("M2b_StatComparison")
ws["A1"] = "M2b StatComparison — Z-score mean vs median cross-check"
ws["A1"].font = Font(size=14, bold=True, color="305496")
ws["A2"] = ("Spec dual-stat: |z_mean| dan |z_median| dievaluasi. flagged_by_mean / "
            "flagged_by_median. Median lebih robust ke outlier (PV3 fault tidak skew median sebanyak mean).")
ws["A2"].font = Font(italic=True, color="595959")

set_header(ws, 4, [
    "PV_string", "R_str",
    "z_mean (vs fleet_mean)", "z_median (vs fleet_median)",
    "|z_mean| > 2.5?", "|z_median| > 2.5?",
    "agree (both flagged)?", "diff |z_mean - z_median|"
])

for i, pv_n in enumerate(range(1, 6)):
    r = 5 + i
    pv_metrics_row = pv_data_start + i
    decision_data_row = decision_row + 2 + i
    ws.cell(row=r, column=1, value=f"PV{pv_n}").border = BORDER
    ws.cell(row=r, column=2,
            value=f"=M2b_PeerZScore!B{pv_metrics_row}").border = BORDER
    ws.cell(row=r, column=2).number_format = "0.0"
    ws.cell(row=r, column=3,
            value=f"=M2b_PeerZScore!B{decision_data_row}").border = BORDER
    ws.cell(row=r, column=3).number_format = "0.00"
    ws.cell(row=r, column=4,
            value=f"=M2b_PeerZScore!C{decision_data_row}").border = BORDER
    ws.cell(row=r, column=4).number_format = "0.00"
    ws.cell(row=r, column=5,
            value=f"=IF(ABS(C{r})>cfg_z_threshold,1,0)").border = BORDER
    ws.cell(row=r, column=6,
            value=f"=IF(ABS(D{r})>cfg_z_threshold,1,0)").border = BORDER
    ws.cell(row=r, column=7,
            value=f"=IF(AND(E{r}=1,F{r}=1),1,IF(OR(E{r}=1,F{r}=1),\"disagree\",0))").border = BORDER
    ws.cell(row=r, column=8,
            value=f"=IFERROR(ABS(C{r}-D{r}),\"\")").border = BORDER
    ws.cell(row=r, column=8).number_format = "0.00"

for letter, w in zip("ABCDEFGH", [12, 11, 22, 22, 16, 17, 19, 22]):
    ws.column_dimensions[letter].width = w


# ===========================================================================
# Sheet: Hampel_Preprocessing — A/B comparison
# ===========================================================================
ws = wb.create_sheet("Hampel_Preprocessing")
ws["A1"] = "Hampel Preprocessing — A/B comparison V series original vs Hampel-cleaned"
ws["A1"].font = Font(size=14, bold=True, color="305496")
ws["A2"] = ("Demo Wave 9: rolling median window 15 (75-min) + MAD-based outlier mask. "
            "Approximation of pvanalytics.quality.outliers.hampel(). Toggle via cfg_enabled.")
ws["A2"].font = Font(italic=True, color="595959")

ws["A4"] = "Note: window 15 butuh ≥15 samples. Dummy 24 timestep cukup untuk demo."
ws["A4"].font = Font(italic=True, size=9, color="A0A0A0")

set_header(ws, 5, [
    "Start Time", "PV3_V_original",
    "rolling_median (window 15)", "abs_dev_from_median",
    "rolling_MAD", "deviation_in_MAD_sigma", "is_outlier", "PV3_V_cleaned",
])

# We'll use PV3 (the fault PV) as the V series under test.
# Note: with only 24 timesteps and window 15, MIN_PERIODS will only kick in for rows 19+.
for ri in range(6, 6 + 24):
    helper_row = 5 + (ri - 6)  # Helpers_M2b row
    ws.cell(row=ri, column=1, value=f"=Helpers_M2b!B{helper_row}").border = BORDER
    ws.cell(row=ri, column=1).number_format = "yyyy-mm-dd hh:mm:ss"

    # PV3_V original (Helpers_M2b col O = PV3_V which is col index 15)
    ws.cell(row=ri, column=2,
            value=f"=Helpers_M2b!O{helper_row}").border = BORDER
    ws.cell(row=ri, column=2).number_format = "0.00"

    # rolling_median window 15 — using OFFSET to grab 15 prior cells.
    # For early rows where < 15 samples available, use available samples.
    # Approximation: =MEDIAN(window_range)
    # window: max(start, ri-14) .. ri (so up to 15 rows lookback)
    win_start = max(6, ri - 14)
    ws.cell(row=ri, column=3,
            value=f"=MEDIAN(B{win_start}:B{ri})").border = BORDER
    ws.cell(row=ri, column=3).number_format = "0.00"

    # abs deviation
    ws.cell(row=ri, column=4,
            value=f"=ABS(B{ri}-C{ri})").border = BORDER
    ws.cell(row=ri, column=4).number_format = "0.00"

    # rolling MAD (median of absolute deviations from median, in window)
    # Approximation: MEDIAN of |B-C| values in window. Use array formula or simpler.
    ws.cell(row=ri, column=5,
            value=f"=MEDIAN(IF(B{win_start}:B{ri}>0,ABS(B{win_start}:B{ri}-C{ri})))").border = BORDER
    ws.cell(row=ri, column=5).number_format = "0.00"

    # deviation_in_MAD_sigma = abs_dev / (1.4826 * MAD) — constant for normal dist
    ws.cell(row=ri, column=6,
            value=f"=IFERROR(D{ri}/(1.4826*E{ri}),0)").border = BORDER
    ws.cell(row=ri, column=6).number_format = "0.00"

    # is_outlier = (sigma > max_deviation)
    ws.cell(row=ri, column=7,
            value=f"=IF(F{ri}>cfg_hampel_max_deviation,1,0)").border = BORDER

    # cleaned value = NaN (empty) if outlier else original
    ws.cell(row=ri, column=8,
            value=f'=IF(G{ri}=1,"",B{ri})').border = BORDER
    ws.cell(row=ri, column=8).number_format = "0.00"

for letter, w in zip("ABCDEFGH", [20, 15, 22, 18, 13, 22, 13, 17]):
    ws.column_dimensions[letter].width = w


# ===========================================================================
# Update Findings_Summary — append M2b high_R findings
# ===========================================================================
ws = wb["Findings_Summary"]
# Find last row with content
last_row = 5
while ws.cell(row=last_row, column=1).value:
    last_row += 1

# Append separator note
ws.cell(row=last_row, column=1, value="--- M2b PeerZScore ---")
ws.cell(row=last_row, column=1).font = SUBHEADER_FONT
ws.cell(row=last_row, column=1).fill = SUBHEADER_FILL
last_row += 1

# Add 5 PV findings (only PV1..PV5)
for i in range(5):
    pv_metrics_row = pv_data_start + i
    decision_data_row = decision_row + 2 + i
    pv_n = i + 1

    ws.cell(row=last_row, column=1, value="M2b_PeerZScore!C").border = BORDER
    ws.cell(row=last_row, column=2, value="M2b_peer_zscore").border = BORDER
    ws.cell(row=last_row, column=3, value="WB05-INV05").border = BORDER
    ws.cell(row=last_row, column=4, value=f"PV{pv_n}").border = BORDER
    if pv_n == 5:
        ws.cell(row=last_row, column=5, value="EMPTY").border = BORDER
        ws.cell(row=last_row, column=6, value="-").border = BORDER
        ws.cell(row=last_row, column=7, value="-").border = BORDER
        ws.cell(row=last_row, column=8, value="-").border = BORDER
    else:
        ws.cell(row=last_row, column=5,
                value=f"=M2b_PeerZScore!I{decision_data_row}").border = BORDER
        ws.cell(row=last_row, column=6,
                value=f"=M2b_PeerZScore!D{decision_data_row}").border = BORDER
        ws.cell(row=last_row, column=6).number_format = "0.00"
        ws.cell(row=last_row, column=7,
                value=f"=cfg_z_threshold").border = BORDER
        ws.cell(row=last_row, column=8,
                value=f"=M2b_PeerZScore!K{decision_data_row}").border = BORDER
    last_row += 1

# Update auto_filter range
ws.auto_filter.ref = f"A4:H{last_row-1}"
sev_range_f = f"E5:E{last_row-1}"
for sev_name, fill in SEV_FILL.items():
    ws.conditional_formatting.add(
        sev_range_f,
        CellIsRule(operator="equal", formula=[f'"{sev_name}"'], fill=fill),
    )


# ===========================================================================
# Save
# ===========================================================================
wb.save(INPUT)
print(f"Workbook updated: {INPUT}")
print(f"Final sheets: {wb.sheetnames}")
