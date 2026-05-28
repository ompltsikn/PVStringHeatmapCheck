"""Build M2 PV Performance Workbook — Iterasi 2: M2eAvailability.

Output: docs/M2_PV_Performance_Workbook.xlsx

Workbook structure (akan tumbuh per iterasi):
  1. README              — petunjuk + log iterasi
  2. Config              — semua threshold + named ranges
  3. Raw_Data            — dummy data realistis (2 inverter × 12 timestep)
  4. EmptyPVMap          — daftar PV slot kosong by design
  5. Helpers_M2e         — status classification + sibling median + qualified mask
  6. M2e_Availability    — uptime per inverter + per-string + severity
  7. M2e_AllStrings      — tabel status setiap (inverter, PV) — replika sheet output Python
  8. Findings_Summary    — agregat semua finding non-NORMAL (saat ini hanya M2e; akan tumbuh)

Iterasi berikutnya akan APPEND sheet baru (Helpers_M2b, M2b_PeerZScore, dll.)
TANPA mengubah sheet yang sudah ada (kecuali Findings_Summary akan
diperluas formula-nya untuk mencakup detector baru).
"""
from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

OUTPUT = Path(__file__).parent / "M2_PV_Performance_Workbook.xlsx"

# --- Styling helpers --------------------------------------------------------
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
SUBHEADER_FONT = Font(name="Calibri", size=10, bold=True)
SEV_FILL = {
    "CRITICAL": PatternFill("solid", fgColor="E06666"),
    "HIGH":     PatternFill("solid", fgColor="F6B26B"),
    "MEDIUM":   PatternFill("solid", fgColor="FFD966"),
    "INFO":     PatternFill("solid", fgColor="A4C2F4"),
    "NORMAL":   PatternFill("solid", fgColor="B6D7A8"),
    "EMPTY":    PatternFill("solid", fgColor="DDDDDD"),
    "NO_DAYLIGHT": PatternFill("solid", fgColor="EFEFEF"),
}

def set_header(ws, row, headers):
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

def auto_width(ws, min_w=10, max_w=42):
    for col in ws.columns:
        try:
            letter = col[0].column_letter
        except AttributeError:
            continue
        max_len = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            try:
                s = str(v)
                # ignore long formulas for width
                if s.startswith("=") and len(s) > 30:
                    continue
                max_len = max(max_len, len(s))
            except Exception:
                pass
        ws.column_dimensions[letter].width = max(min_w, min(max_w, max_len + 2))


# ===========================================================================
wb = Workbook()
# Remove default sheet
default = wb.active
wb.remove(default)


# ===========================================================================
# Sheet 1: README
# ===========================================================================
ws = wb.create_sheet("README")
ws["A1"] = "M2 PV Performance — Workbook Excel"
ws["A1"].font = Font(size=16, bold=True, color="305496")
ws["A2"] = "Reverse engineering dari pv_pipeline/ (notebook 20260514stringmap_v1.5.ipynb)"
ws["A2"].font = Font(italic=True, color="595959")

row = 4
ws[f"A{row}"] = "Iterasi log"
ws[f"A{row}"].font = SUBHEADER_FONT
row += 1
log_data = [
    ("Iterasi", "Tanggal", "Detector", "Sheet yang ditambah"),
    ("1 (Phase 1)", "2026-05-28", "—", "Hanya dokumen overview (.docx + .md). Belum ada Excel."),
    ("2", "2026-05-28", "M2eAvailability", "README, Config, Raw_Data, EmptyPVMap, Helpers_M2e, M2e_Availability, M2e_AllStrings, Findings_Summary"),
]
for i, rowdata in enumerate(log_data):
    for j, v in enumerate(rowdata):
        c = ws.cell(row=row+i, column=1+j, value=v)
        if i == 0:
            c.fill = SUBHEADER_FILL
            c.font = SUBHEADER_FONT
        c.border = BORDER

row += len(log_data) + 2
ws[f"A{row}"] = "Cara membaca"
ws[f"A{row}"].font = SUBHEADER_FONT
notes = [
    "Sheet 'Config' adalah single source of truth untuk semua threshold. Edit di sini → seluruh workbook recalc.",
    "Sheet 'Raw_Data' berisi data dummy realistis. Ganti dengan data aktual mu (paste over) untuk pakai workbook ini di lapangan.",
    "Sheet 'Helpers_M2e' adalah JEMBATAN antara Raw_Data dan M2e_Availability. Tiap kolom punya formula yang ekuivalen dengan satu langkah di pv_pipeline/availability.py.",
    "Sheet 'M2e_Availability' menampilkan hasil akhir per inverter & per string + severity. Sheet 'M2e_AllStrings' mereplika output sheet xlsx Python.",
    "Severity dihitung via IFS() dari uptime_pct ke threshold di Config.",
    "Workbook ini akan TUMBUH per iterasi — sheet baru ditambah tanpa mengubah yang lama.",
]
for i, n in enumerate(notes):
    c = ws.cell(row=row+1+i, column=1, value=f"• {n}")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row+1+i].height = 30
ws.column_dimensions["A"].width = 110


# ===========================================================================
# Sheet 2: Config
# ===========================================================================
ws = wb.create_sheet("Config")
ws["A1"] = "M2 Configuration (mirror config/m2_config.yaml)"
ws["A1"].font = Font(size=14, bold=True, color="305496")

row = 3
ws[f"A{row}"] = "Section"
ws[f"B{row}"] = "Key"
ws[f"C{row}"] = "Value"
ws[f"D{row}"] = "Unit / Notes"
for col in "ABCD":
    c = ws[f"{col}{row}"]
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.border = BORDER

config_rows = [
    # M2e — Severity thresholds
    ("m2e.severity", "critical_below",     90.0, "% uptime; < ini → CRITICAL"),
    ("m2e.severity", "high_below",         95.0, "% uptime; < ini → HIGH"),
    ("m2e.severity", "medium_below",       97.0, "% uptime; < ini → MEDIUM"),
    ("m2e.severity", "info_below",         99.0, "% uptime; < ini → INFO; ≥ ini → NORMAL"),
    ("m2e.severity", "emit_normal",        0,    "1=emit NORMAL findings, 0=skip"),
    # M2e — String proxy
    ("m2e.string_proxy", "pstr_zero_threshold_kw",     0.1,  "kW; PV power < ini → string dianggap mati"),
    ("m2e.string_proxy", "sibling_median_active_kw",   1.0,  "kW; median sibling >= ini → timestamp 'produktif'"),
    ("m2e.string_proxy", "min_active_siblings_pct",    50.0, "%; minimum % sibling aktif → 'qualified' timestamp"),
    ("m2e.string_proxy", "debounce_consecutive_steps", 2,    "step; run candidate >= ini → genuine event. DUMMY=2 (original spec). PRODUKSI IKN=20 (~1h40m persist)."),
    # Interval estimate
    ("system", "interval_minutes",         5,    "menit; median selisih Start Time (auto-detected, override saat dummy)"),
    # PV inventory
    ("system", "pv_max",                   28,   "PV string per inverter (hardcoded Huawei cap)"),
    ("system", "n_pv_total_in_workbook",   5,    "PV string yang ada di sheet ini (PV1..PV5)"),
]

named = {}  # nama → cell ref
for i, (section, key, value, note) in enumerate(config_rows, start=row + 1):
    ws.cell(row=i, column=1, value=section).border = BORDER
    ws.cell(row=i, column=2, value=key).border = BORDER
    c = ws.cell(row=i, column=3, value=value)
    c.border = BORDER
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="center")
    ws.cell(row=i, column=4, value=note).border = BORDER
    name = f"cfg_{key}"
    named[name] = f"Config!$C${i}"

# Define named ranges so formulas in other sheets can reference cfg_xxx
for name, ref in named.items():
    dn = DefinedName(name, attr_text=ref)
    wb.defined_names[name] = dn

ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 32
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 60


# ===========================================================================
# Sheet 3: Raw_Data — dummy data
# ===========================================================================
# Scenario:
#   2 inverters × 12 timestep (5-min interval, 10:00 → 10:55), 5 PV strings each.
#   WB02-INV01: all 12 ON; PV3 = 0 kW from timestep 3..8 (6 consecutive steps,
#                debounce 20 NOT met if we use default 20; we override to 2 in Config
#                for demo to show fired event).
#   WB02-INV02: 8 ON + 4 DOWN (timestep 4..7) → inverter-level uptime 8/12 = 66.67% CRITICAL.
#   PV5 di empty_pv_map untuk DEMO EMPTY logic (we add WB02-INV01 PV5).
#
# Penting: untuk dummy 12 timestep, default debounce=20 (1h40m) tidak akan
# fire. Saya OVERRIDE debounce di sheet Config = 2 (sesuai DEFAULT lama
# 2026-05-08, sebelum tuning 20). Notes: di lapangan IKN dipakai 20.

ws = wb.create_sheet("Raw_Data")
ws["A1"] = "Raw_Data (dummy realistis — 2 inverter × 12 timestep)"
ws["A1"].font = Font(size=14, bold=True, color="305496")
ws["A2"] = ("Skenario: INV01 punya PV3 fault 6-timestep; INV02 down 4-timestep. "
            "PV5 (kedua inverter) ada di EmptyPVMap.")
ws["A2"].font = Font(italic=True, color="595959")

headers = ["Inverter_ID", "Start Time", "Inverter status",
           "PV1 Power(kW)", "PV2 Power(kW)", "PV3 Power(kW)",
           "PV4 Power(kW)", "PV5 Power(kW)"]
set_header(ws, 4, headers)

start_time = datetime(2026, 5, 14, 10, 0, 0)
# Build INV01 rows (all ON, PV3 fault from t=3..8)
inv01_rows = []
for t in range(12):
    ts = start_time + timedelta(minutes=5 * t)
    pv1 = 5.0 + 0.05 * (t % 4)
    pv2 = 4.9 + 0.04 * (t % 5)
    pv3 = 5.1 if t < 3 or t > 8 else 0.0
    pv4 = 5.0 + 0.03 * (t % 3)
    pv5 = 0.0  # EMPTY slot
    inv01_rows.append(("WB02-INV01", ts, "Grid connected", pv1, pv2, pv3, pv4, pv5))
# Build INV02 rows (ON for t=0..3, DOWN for t=4..7, ON for t=8..11)
inv02_rows = []
for t in range(12):
    ts = start_time + timedelta(minutes=5 * t)
    if 4 <= t <= 7:
        status = "Shutdown command"
        pv1 = pv2 = pv3 = pv4 = pv5 = 0.0
    else:
        status = "Grid connected"
        pv1 = 5.0
        pv2 = 5.0
        pv3 = 5.0
        pv4 = 5.0
        pv5 = 0.0
    inv02_rows.append(("WB02-INV02", ts, status, pv1, pv2, pv3, pv4, pv5))

all_rows = inv01_rows + inv02_rows  # 24 rows total

for ri, row_tuple in enumerate(all_rows, start=5):
    for ci, v in enumerate(row_tuple, start=1):
        c = ws.cell(row=ri, column=ci, value=v)
        c.border = BORDER
        if ci == 2:  # timestamp
            c.number_format = "yyyy-mm-dd hh:mm:ss"
        elif ci >= 4:  # PV power kW
            c.number_format = "0.000"

ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 22
for col_letter in ["D", "E", "F", "G", "H"]:
    ws.column_dimensions[col_letter].width = 14

LAST_RAW = 4 + len(all_rows)  # = 28


# ===========================================================================
# Sheet 4: EmptyPVMap
# ===========================================================================
ws = wb.create_sheet("EmptyPVMap")
ws["A1"] = "EmptyPVMap — slot PV yang fisik tidak terpasang (skip dari peer comparison)"
ws["A1"].font = Font(size=14, bold=True, color="305496")
ws["A2"] = "Replica config/strings.yaml (245 inverter dalam file aktual). Di sini hanya entry yang relevan untuk dummy."
ws["A2"].font = Font(italic=True, color="595959")

set_header(ws, 4, ["Inverter_ID", "Empty PV indices (CSV)", "Note"])
empty_map_rows = [
    ("WB02-INV01", "5",                  "DEMO: PV5 EMPTY by design"),
    ("WB02-INV02", "5",                  "DEMO: PV5 EMPTY by design"),
    ("WB01-INV01", "19,20,21,22,23,24,25,26,27,28", "Pola WB01 produksi (PV19-28 kosong)"),
    ("WB03-INV01", "",                                "(contoh: tidak ada slot kosong)"),
]
for ri, (inv, idxs, note) in enumerate(empty_map_rows, start=5):
    ws.cell(row=ri, column=1, value=inv).border = BORDER
    ws.cell(row=ri, column=2, value=idxs).border = BORDER
    ws.cell(row=ri, column=3, value=note).border = BORDER

ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 32
ws.column_dimensions["C"].width = 60


# ===========================================================================
# Sheet 5: Helpers_M2e
# ===========================================================================
# Per-row helper kolom (mengekstrak logika dari availability.py):
#   - _status_class    : ON / DOWN / TRANSITIONAL / UNKNOWN (via formula klasifikasi)
#   - sib_median_kW    : MEDIAN(PV1..PV5) per row (skipna)
#   - active_count     : COUNT PVx > pstr_zero_threshold
#   - active_pct       : 100 × active_count / n_pv_total
#   - qualified        : (sib_median >= sib_med_th) AND (active_pct >= min_act_pct)
#   - is_on            : 1 jika _status_class = "ON"
#   - is_daylight_ts   : 1 jika ada minimal 1 inverter ON di timestamp ini (proxy daylight)
#
# Kemudian per-PV mask + debounce:
#   - cand_PV<n>       : 1 jika (qualified AND PV<n> < pstr_zero AND inverter ON)
#   - run_id_PV<n>     : counter run (increments on 0→1 transition per inverter group)
#   - run_len_PV<n>    : COUNTIFS panjang run saat ini (only count cand=1 dengan run_id sama)
#   - in_event_PV<n>   : 1 jika cand=1 AND run_len >= debounce
ws = wb.create_sheet("Helpers_M2e")
ws["A1"] = "Helpers_M2e — derived columns dari Raw_Data (per-row formulas)"
ws["A1"].font = Font(size=14, bold=True, color="305496")
ws["A2"] = ("Setiap kolom = 1 langkah di pv_pipeline/availability.py. Edit Raw_Data → recalc otomatis.")
ws["A2"].font = Font(italic=True, color="595959")

# Headers
helper_headers = [
    "Inverter_ID", "Start Time", "status_raw",         # 1-3 mirror raw
    "_status_class",                                    # 4
    "sib_median_kW", "active_count", "active_pct",     # 5-7
    "qualified",                                        # 8
    "is_on",                                            # 9
    # Per-PV columns (5 PV × 4 helpers = 20 cols)
    "cand_PV1", "run_id_PV1", "run_len_PV1", "in_event_PV1",   # 10-13
    "cand_PV2", "run_id_PV2", "run_len_PV2", "in_event_PV2",   # 14-17
    "cand_PV3", "run_id_PV3", "run_len_PV3", "in_event_PV3",   # 18-21
    "cand_PV4", "run_id_PV4", "run_len_PV4", "in_event_PV4",   # 22-25
    "cand_PV5", "run_id_PV5", "run_len_PV5", "in_event_PV5",   # 26-29
]
set_header(ws, 4, helper_headers)
# Freeze panes
ws.freeze_panes = "C5"

# Per-row formulas. Raw_Data rows start at row 5, end at row LAST_RAW (28).
for ri in range(5, LAST_RAW + 1):
    raw = ri  # same row index since both start at row 5

    # 1-3 mirror raw (cell references)
    ws.cell(row=ri, column=1, value=f"=Raw_Data!A{raw}").border = BORDER
    c = ws.cell(row=ri, column=2, value=f"=Raw_Data!B{raw}")
    c.border = BORDER
    c.number_format = "yyyy-mm-dd hh:mm:ss"
    ws.cell(row=ri, column=3, value=f"=Raw_Data!C{raw}").border = BORDER

    # 4: _status_class
    # Priority: down > on > transitional. Hardcoded keyword list for dummy clarity.
    # Real Python uses config-driven substring match; here we replicate the 3 most
    # important keywords (shutdown/fault → DOWN; grid connected → ON; standby/no sunlight → TRANSITIONAL).
    status_formula = (
        f'=IF(ISNUMBER(SEARCH("shutdown",LOWER(C{ri})))+ISNUMBER(SEARCH("fault",LOWER(C{ri}))),"DOWN",'
        f'IF(ISNUMBER(SEARCH("grid connected",LOWER(C{ri})))+ISNUMBER(SEARCH("on-grid",LOWER(C{ri}))),"ON",'
        f'IF(ISNUMBER(SEARCH("standby",LOWER(C{ri})))+ISNUMBER(SEARCH("no sunlight",LOWER(C{ri}))),"TRANSITIONAL",'
        f'"UNKNOWN")))'
    )
    ws.cell(row=ri, column=4, value=status_formula).border = BORDER

    # 5: sib_median_kW = MEDIAN PV1..PV5 (referensi ke Raw_Data row)
    # In real code: skipna across all PV cols. Excel MEDIAN ignores blanks but
    # treats 0 as value. Dummy data: 0 means down/empty. For EMPTY slot we'd
    # want to exclude, but pure Python median includes 0. Match Python behavior:
    # use MEDIAN on D..H including 0.
    ws.cell(row=ri, column=5,
            value=f"=MEDIAN(Raw_Data!D{raw}:H{raw})").border = BORDER
    ws.cell(row=ri, column=5).number_format = "0.000"

    # 6: active_count = COUNTIF(PVx > pstr_zero)
    ws.cell(row=ri, column=6,
            value=f"=COUNTIF(Raw_Data!D{raw}:H{raw}, \">\"&cfg_pstr_zero_threshold_kw)").border = BORDER

    # 7: active_pct = 100 * active_count / n_pv_total
    ws.cell(row=ri, column=7,
            value=f"=100*F{ri}/cfg_n_pv_total_in_workbook").border = BORDER
    ws.cell(row=ri, column=7).number_format = "0.0"

    # 8: qualified = (sib_median >= sib_med_th) AND (active_pct >= min_act_pct)
    ws.cell(row=ri, column=8,
            value=f"=IF(AND(E{ri}>=cfg_sibling_median_active_kw, G{ri}>=cfg_min_active_siblings_pct),1,0)").border = BORDER

    # 9: is_on = (_status_class = "ON")
    ws.cell(row=ri, column=9,
            value=f"=IF(D{ri}=\"ON\",1,0)").border = BORDER

    # Per-PV (PV1..PV5): cand, run_id, run_len, in_event
    # Group: per inverter, per PV column.
    # cand_PV<n> = (is_on AND qualified AND PV<n> < pstr_zero)
    # run_id_PV<n>: cumulative count of 0→1 transitions WITHIN current Inverter_ID
    #   = IF(this row Inverter != prev row Inverter, IF(cand=1, 1, 0),
    #         IF(AND(cand=1, prev_cand=0), prev_run_id + 1,
    #            IF(cand=1, prev_run_id, prev_run_id)))
    # run_len_PV<n> = COUNTIFS(inv range, this inv, run_id range, this run_id, cand range, 1)
    # in_event_PV<n> = (cand=1 AND run_len >= debounce)
    pv_cols_letters = {1: "D", 2: "E", 3: "F", 4: "G", 5: "H"}  # in Raw_Data
    cand_col_letters = {1: "J", 2: "N", 3: "R", 4: "V", 5: "Z"}
    runid_col_letters = {1: "K", 2: "O", 3: "S", 4: "W", 5: "AA"}
    runlen_col_letters = {1: "L", 2: "P", 3: "T", 4: "X", 5: "AB"}
    inevent_col_letters = {1: "M", 2: "Q", 3: "U", 4: "Y", 5: "AC"}

    for pv_n, pv_letter_raw in pv_cols_letters.items():
        cand_col = cand_col_letters[pv_n]
        runid_col = runid_col_letters[pv_n]
        runlen_col = runlen_col_letters[pv_n]
        inevent_col = inevent_col_letters[pv_n]

        # cand_PV<n>
        cand_formula = (
            f"=IF(AND(I{ri}=1, H{ri}=1, Raw_Data!{pv_letter_raw}{raw}<cfg_pstr_zero_threshold_kw),1,0)"
        )
        ws.cell(row=ri, column=openpyxl.utils.column_index_from_string(cand_col),
                value=cand_formula).border = BORDER

        # run_id_PV<n>
        if ri == 5:
            # first row: run_id = cand (1 if cand=1, else 0)
            runid_formula = f"=IF({cand_col}{ri}=1,1,0)"
        else:
            # If new inverter group (Inverter_ID different from prev row), reset based on cand
            # Else: if cand transitions 0→1, increment; else keep previous; if cand=0, keep prev
            runid_formula = (
                f"=IF(A{ri}<>A{ri-1},"
                f"IF({cand_col}{ri}=1, MAX(${runid_col}$5:{runid_col}{ri-1})+1, 0),"  # new inverter
                f"IF(AND({cand_col}{ri}=1, {cand_col}{ri-1}=0), {runid_col}{ri-1}+1,"
                f"IF({cand_col}{ri}=1, {runid_col}{ri-1}, {runid_col}{ri-1})))"
            )
        ws.cell(row=ri, column=openpyxl.utils.column_index_from_string(runid_col),
                value=runid_formula).border = BORDER

        # run_len_PV<n> = COUNTIFS within same inverter, same run_id, cand=1
        runlen_formula = (
            f"=IF({cand_col}{ri}=1,"
            f"COUNTIFS($A$5:$A${LAST_RAW},A{ri},"
            f"${runid_col}$5:${runid_col}${LAST_RAW},{runid_col}{ri},"
            f"${cand_col}$5:${cand_col}${LAST_RAW},1),0)"
        )
        ws.cell(row=ri, column=openpyxl.utils.column_index_from_string(runlen_col),
                value=runlen_formula).border = BORDER

        # in_event_PV<n>
        inevent_formula = (
            f"=IF(AND({cand_col}{ri}=1, {runlen_col}{ri}>=cfg_debounce_consecutive_steps),1,0)"
        )
        ws.cell(row=ri, column=openpyxl.utils.column_index_from_string(inevent_col),
                value=inevent_formula).border = BORDER

# Column widths
ws.column_dimensions["A"].width = 13
ws.column_dimensions["B"].width = 19
ws.column_dimensions["C"].width = 18
for col_letter in ["D", "E", "F", "G", "H", "I"]:
    ws.column_dimensions[col_letter].width = 13
# Per-PV groups
for letter in ["J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z","AA","AB","AC"]:
    ws.column_dimensions[letter].width = 10


# ===========================================================================
# Sheet 6: M2e_Availability
# ===========================================================================
ws = wb.create_sheet("M2e_Availability")
ws["A1"] = "M2e Availability — output finding per inverter & per string"
ws["A1"].font = Font(size=14, bold=True, color="305496")
ws["A2"] = ("Section A: per-inverter uptime%. Section B: per-string proxy down. "
            "Severity di-derive dari uptime_pct vs threshold di Config.")
ws["A2"].font = Font(italic=True, color="595959")

# --- Section A: Per-inverter uptime ---
ws["A4"] = "A. Per-inverter daily uptime"
ws["A4"].font = SUBHEADER_FONT
ws["A4"].fill = SUBHEADER_FILL

set_header(ws, 5, ["Inverter_ID", "n_on", "n_down", "denom", "uptime_pct",
                   "downtime_min", "severity", "threshold_breached", "message"])

# Hardcode inverter IDs (sesuai dummy): WB02-INV01, WB02-INV02
inv_ids = ["WB02-INV01", "WB02-INV02"]
inv_start_row = 6
helper_inv_range = f"Helpers_M2e!$A$5:$A${LAST_RAW}"
helper_status_range = f"Helpers_M2e!$D$5:$D${LAST_RAW}"

for i, inv in enumerate(inv_ids):
    r = inv_start_row + i
    ws.cell(row=r, column=1, value=inv).border = BORDER

    # n_on = COUNTIFS(inv_range, this_inv, status_range, "ON")
    ws.cell(row=r, column=2,
            value=f'=COUNTIFS({helper_inv_range}, A{r}, {helper_status_range}, "ON")').border = BORDER
    # n_down
    ws.cell(row=r, column=3,
            value=f'=COUNTIFS({helper_inv_range}, A{r}, {helper_status_range}, "DOWN")').border = BORDER
    # denom = n_on + n_down
    ws.cell(row=r, column=4, value=f"=B{r}+C{r}").border = BORDER
    # uptime_pct = 100 * n_on / denom (NaN-safe via IF)
    ws.cell(row=r, column=5,
            value=f'=IF(D{r}=0, NA(), 100*B{r}/D{r})').border = BORDER
    ws.cell(row=r, column=5).number_format = "0.00"
    # downtime_min = n_down * interval
    ws.cell(row=r, column=6,
            value=f'=C{r}*cfg_interval_minutes').border = BORDER
    # severity via nested IF (compatible Excel 2010+, LibreOffice). IFS belum universal.
    ws.cell(row=r, column=7,
            value=(
                f'=IF(ISNA(E{r}),"NO_DATA",'
                f'IF(E{r}<cfg_critical_below,"CRITICAL",'
                f'IF(E{r}<cfg_high_below,"HIGH",'
                f'IF(E{r}<cfg_medium_below,"MEDIUM",'
                f'IF(E{r}<cfg_info_below,"INFO","NORMAL")))))'
            )).border = BORDER
    # threshold_breached: nilai cfg_xx_below yang dilanggar
    ws.cell(row=r, column=8,
            value=(
                f'=IF(ISNA(E{r}),"",'
                f'IF(E{r}<cfg_critical_below,cfg_critical_below,'
                f'IF(E{r}<cfg_high_below,cfg_high_below,'
                f'IF(E{r}<cfg_medium_below,cfg_medium_below,'
                f'IF(E{r}<cfg_info_below,cfg_info_below,cfg_info_below)))))'
            )).border = BORDER
    # message
    ws.cell(row=r, column=9,
            value=f'="inverter uptime "&TEXT(E{r},"0.00")&"% (n_on="&B{r}&", n_down="&C{r}&")"').border = BORDER

# Conditional fill for severity column G
for i in range(len(inv_ids)):
    r = inv_start_row + i
    # cannot do real conditional in openpyxl easily without rules; we just leave styling
    # but we'll add conditional formatting below.

from openpyxl.formatting.rule import CellIsRule, FormulaRule
sev_rules_range = f"G{inv_start_row}:G{inv_start_row + len(inv_ids) - 1}"
for sev_name, fill in SEV_FILL.items():
    ws.conditional_formatting.add(
        sev_rules_range,
        FormulaRule(formula=[f'EXACT(G{inv_start_row},"{sev_name}")'], fill=fill, stopIfTrue=False),
    )
# Hmm formula above is row-fixed; need per-row. Use better approach:
# Re-do per-cell.
ws.conditional_formatting = openpyxl.formatting.formatting.ConditionalFormattingList()
for sev_name, fill in SEV_FILL.items():
    ws.conditional_formatting.add(
        sev_rules_range,
        CellIsRule(operator="equal", formula=[f'"{sev_name}"'], fill=fill),
    )


# --- Section B: Per-string proxy down (Inverter × PV) ---
sec_b_row = inv_start_row + len(inv_ids) + 3
ws.cell(row=sec_b_row, column=1, value="B. Per-string proxy down (event_minutes_total + uptime_pct + severity)").font = SUBHEADER_FONT
ws.cell(row=sec_b_row, column=1).fill = SUBHEADER_FILL

set_header(ws, sec_b_row + 1, ["Inverter_ID", "PV_string",
                                "daylight_minutes_inv", "event_minutes", "n_events",
                                "string_uptime_pct", "severity", "is_empty", "message"])

# Loop per (inverter × PV1..5)
b_data_start = sec_b_row + 2
row_cursor = b_data_start
for inv in inv_ids:
    for pv_n in range(1, 6):
        pv_label = f"PV{pv_n}"
        cand_col = {1:"J",2:"N",3:"R",4:"V",5:"Z"}[pv_n]
        inevent_col = {1:"M",2:"Q",3:"U",4:"Y",5:"AC"}[pv_n]

        ws.cell(row=row_cursor, column=1, value=inv).border = BORDER
        ws.cell(row=row_cursor, column=2, value=pv_label).border = BORDER

        # daylight_minutes_inv = (count of is_on=1 for this inverter) × interval
        ws.cell(row=row_cursor, column=3,
                value=(
                    f'=COUNTIFS({helper_inv_range}, A{row_cursor}, '
                    f'Helpers_M2e!$I$5:$I${LAST_RAW}, 1) * cfg_interval_minutes'
                )).border = BORDER

        # event_minutes = (count in_event_PV<n>=1 for this inverter) × interval
        ws.cell(row=row_cursor, column=4,
                value=(
                    f'=COUNTIFS({helper_inv_range}, A{row_cursor}, '
                    f'Helpers_M2e!${inevent_col}$5:${inevent_col}${LAST_RAW}, 1) * cfg_interval_minutes'
                )).border = BORDER

        # n_events = count distinct run_id values for this inverter with in_event=1
        # Approximated: count of cells where (inverter matches AND cand=1 AND prev_cand=0 AND run_len >= debounce)
        # Simpler proxy: count rows where (in_event=1 AND prev row's in_event=0 OR prev row's inverter different)
        # For simplicity in workbook, use: SUMPRODUCT counting transitions 0→1 in in_event with inverter match
        runid_col = {1:"K",2:"O",3:"S",4:"W",5:"AA"}[pv_n]
        # n_events = number of DISTINCT run_id values where in_event=1 for this inverter
        # Excel doesn't have COUNTUNIQUEIFS pre-365. Use SUMPRODUCT(1/COUNTIFS pattern):
        # Simpler: count of (in_event=1 AND first row in the run for this inverter)
        n_events_formula = (
            f'=SUMPRODUCT(({helper_inv_range}=A{row_cursor})*'
            f'(Helpers_M2e!${inevent_col}$5:${inevent_col}${LAST_RAW}=1)*'
            f'((ROW({helper_inv_range})=5)+'
            f'(Helpers_M2e!${inevent_col}$4:${inevent_col}${LAST_RAW-1}<>1)+'
            f'({helper_inv_range}<>OFFSET({helper_inv_range},-1,0))>=1)'
            f'/MAX(1,SUMPRODUCT(({helper_inv_range}=A{row_cursor})*'
            f'(Helpers_M2e!${inevent_col}$5:${inevent_col}${LAST_RAW}=1))))'
        )
        # The above is messy; replace with simpler proxy: COUNT of transitions
        # in_event[i]=1 AND (i==1 OR in_event[i-1]=0 OR inverter[i]<>inverter[i-1])
        # Using SUMPRODUCT cleanly:
        n_events_simple = (
            f'=SUMPRODUCT('
            f'(Helpers_M2e!${"A"}$5:${"A"}${LAST_RAW}=A{row_cursor})*'
            f'(Helpers_M2e!${inevent_col}$5:${inevent_col}${LAST_RAW}=1)*'
            f'(IF(ROW(Helpers_M2e!${inevent_col}$5:${inevent_col}${LAST_RAW})=5,1,'
            f'(Helpers_M2e!${inevent_col}$4:${inevent_col}${LAST_RAW-1}=0)+'
            f'(Helpers_M2e!${"A"}$5:${"A"}${LAST_RAW}<>Helpers_M2e!${"A"}$4:${"A"}${LAST_RAW-1})>=1)))'
        )
        # Even simpler: use a fallback that counts qualified_runs by counting (in_event[i]=1 AND (in_event[i-1]=0 OR row=first OR inverter changed))
        # Use simplest formula that works (no array madness):
        # n_events = SUMPRODUCT( (inverter_range = A{row}) * (in_event range = 1) * (start_of_run) )
        # start_of_run[i] = 1 if (in_event[i]=1) AND (i=1 OR in_event[i-1]=0 OR inverter[i]<>inverter[i-1])
        # We'll build start_of_run column inline:
        n_events_clean = (
            f'=SUMPRODUCT(({helper_inv_range}=A{row_cursor})*'
            f'(Helpers_M2e!${inevent_col}$5:${inevent_col}${LAST_RAW}=1)*'
            f'((Helpers_M2e!${inevent_col}$4:${inevent_col}${LAST_RAW-1}=0)+'
            f'({helper_inv_range}<>Helpers_M2e!$A$4:$A${LAST_RAW-1})>=1))'
        )
        ws.cell(row=row_cursor, column=5, value=n_events_clean).border = BORDER

        # string_uptime_pct = MAX(0, 100 - 100*event_minutes/daylight_minutes_inv)
        # Handle daylight=0 (NaN-safe)
        ws.cell(row=row_cursor, column=6,
                value=(
                    f'=IF(C{row_cursor}=0, NA(), MAX(0, 100 - 100*D{row_cursor}/C{row_cursor}))'
                )).border = BORDER
        ws.cell(row=row_cursor, column=6).number_format = "0.00"

        # severity (nested IF for max compat)
        ws.cell(row=row_cursor, column=7,
                value=(
                    f'=IF(H{row_cursor}=1,"EMPTY",'
                    f'IF(C{row_cursor}=0,"NO_DAYLIGHT",'
                    f'IF(F{row_cursor}<cfg_critical_below,"CRITICAL",'
                    f'IF(F{row_cursor}<cfg_high_below,"HIGH",'
                    f'IF(F{row_cursor}<cfg_medium_below,"MEDIUM",'
                    f'IF(F{row_cursor}<cfg_info_below,"INFO","NORMAL"))))))'
                )).border = BORDER

        # is_empty: cross-check EmptyPVMap (lookup CSV string and FIND number)
        # For dummy: PV5 is empty for both INV01 & INV02 → mark.
        # Excel: =IF(ISNUMBER(FIND(","&pv_n&",", ","&VLOOKUP(inv, EmptyPVMap, 2, FALSE)&",")), 1, 0)
        is_empty_formula = (
            f'=IFERROR(IF(ISNUMBER(SEARCH('
            f'","&{pv_n}&",", '
            f'","&VLOOKUP(A{row_cursor}, EmptyPVMap!$A$5:$B$8, 2, FALSE)&","'
            f')), 1, 0), 0)'
        )
        ws.cell(row=row_cursor, column=8, value=is_empty_formula).border = BORDER

        # message
        ws.cell(row=row_cursor, column=9,
                value=(
                    f'=IF(H{row_cursor}=1, "EMPTY (skip)", '
                    f'IF(F{row_cursor}>=cfg_info_below, "NORMAL", '
                    f'A{row_cursor}&" "&B{row_cursor}&" proxy-down "&D{row_cursor}&'
                    f'"min/"&C{row_cursor}&"min daylight ("&TEXT(F{row_cursor},"0.00")&"% uptime)"))'
                )).border = BORDER

        row_cursor += 1

# Conditional formatting for severity in Section B
sec_b_sev_range = f"G{b_data_start}:G{row_cursor-1}"
for sev_name, fill in SEV_FILL.items():
    ws.conditional_formatting.add(
        sec_b_sev_range,
        CellIsRule(operator="equal", formula=[f'"{sev_name}"'], fill=fill),
    )

ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 10
ws.column_dimensions["C"].width = 19
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 10
ws.column_dimensions["F"].width = 16
ws.column_dimensions["G"].width = 14
ws.column_dimensions["H"].width = 9
ws.column_dimensions["I"].width = 70
ws.freeze_panes = "A6"


# ===========================================================================
# Sheet 7: M2e_AllStrings (replika output sheet xlsx Python)
# ===========================================================================
ws = wb.create_sheet("M2e_AllStrings")
ws["A1"] = "M2e_AllStrings — tabel rapi per (inverter × PV string) — replika output Python"
ws["A1"].font = Font(size=14, bold=True, color="305496")
ws["A2"] = ("Setiap baris = 1 PV slot. status ∈ {EMPTY, NO_DAYLIGHT, NORMAL, INFO, MEDIUM, HIGH, CRITICAL}. "
            "Pull dari M2e_Availability Section B.")
ws["A2"].font = Font(italic=True, color="595959")

set_header(ws, 4, ["inverter_id", "pv_string", "status",
                   "uptime_pct", "downtime_minutes", "event_minutes",
                   "n_events", "daylight_minutes"])

# Pull data from M2e_Availability Section B (which is at rows b_data_start..row_cursor-1)
all_strings_start = b_data_start
all_strings_end = row_cursor - 1

out_row = 5
for src_row in range(all_strings_start, all_strings_end + 1):
    ws.cell(row=out_row, column=1, value=f"=M2e_Availability!A{src_row}").border = BORDER
    ws.cell(row=out_row, column=2, value=f"=M2e_Availability!B{src_row}").border = BORDER
    ws.cell(row=out_row, column=3, value=f"=M2e_Availability!G{src_row}").border = BORDER
    # uptime
    ws.cell(row=out_row, column=4,
            value=f'=IF(OR(M2e_Availability!G{src_row}="EMPTY",M2e_Availability!G{src_row}="NO_DAYLIGHT"),"",M2e_Availability!F{src_row})').border = BORDER
    ws.cell(row=out_row, column=4).number_format = "0.00"
    # downtime_minutes == event_minutes in Python output
    ws.cell(row=out_row, column=5,
            value=f'=IF(OR(M2e_Availability!G{src_row}="EMPTY",M2e_Availability!G{src_row}="NO_DAYLIGHT"),"",M2e_Availability!D{src_row})').border = BORDER
    ws.cell(row=out_row, column=6,
            value=f'=IF(OR(M2e_Availability!G{src_row}="EMPTY",M2e_Availability!G{src_row}="NO_DAYLIGHT"),"",M2e_Availability!D{src_row})').border = BORDER
    ws.cell(row=out_row, column=7,
            value=f'=IF(M2e_Availability!G{src_row}="EMPTY",0,M2e_Availability!E{src_row})').border = BORDER
    ws.cell(row=out_row, column=8, value=f'=M2e_Availability!C{src_row}').border = BORDER
    out_row += 1

# Conditional formatting per status
sev_range = f"C5:C{out_row-1}"
for sev_name, fill in SEV_FILL.items():
    ws.conditional_formatting.add(
        sev_range,
        CellIsRule(operator="equal", formula=[f'"{sev_name}"'], fill=fill),
    )

for letter, w in zip("ABCDEFGH", [14, 10, 14, 12, 16, 14, 10, 17]):
    ws.column_dimensions[letter].width = w
ws.freeze_panes = "A5"


# ===========================================================================
# Sheet 8: Findings_Summary
# ===========================================================================
ws = wb.create_sheet("Findings_Summary")
ws["A1"] = "Findings_Summary — agregat semua finding NON-NORMAL & NON-EMPTY"
ws["A1"].font = Font(size=14, bold=True, color="305496")
ws["A2"] = ("Saat ini hanya M2e. Akan tumbuh: M2b, M2a, IsolationForest, dst. "
            "Setiap iterasi tambah sheet detector → tambah baris di sini.")
ws["A2"].font = Font(italic=True, color="595959")

set_header(ws, 4, ["Source Sheet", "sub_module", "inverter_id", "pv_string",
                   "severity", "value (uptime_pct)", "threshold", "message"])

# Pull from M2e_Availability Section A (inverter findings) and Section B (string findings)
# Filter out NORMAL & NO_DATA & EMPTY & NO_DAYLIGHT
# We can't easily filter in formulas without Excel 365 FILTER(); we'll use static rows
# that reference and let user manually filter or apply AutoFilter.

# Section A rows: 6..6+len(inv_ids)-1
row_out = 5
for src_row in range(inv_start_row, inv_start_row + len(inv_ids)):
    ws.cell(row=row_out, column=1, value="M2e_Availability!A (inverter)").border = BORDER
    ws.cell(row=row_out, column=2, value="M2e_inverter").border = BORDER
    ws.cell(row=row_out, column=3, value=f"=M2e_Availability!A{src_row}").border = BORDER
    ws.cell(row=row_out, column=4, value="").border = BORDER
    ws.cell(row=row_out, column=5, value=f"=M2e_Availability!G{src_row}").border = BORDER
    ws.cell(row=row_out, column=6, value=f"=M2e_Availability!E{src_row}").border = BORDER
    ws.cell(row=row_out, column=6).number_format = "0.00"
    ws.cell(row=row_out, column=7, value=f"=M2e_Availability!H{src_row}").border = BORDER
    ws.cell(row=row_out, column=8, value=f"=M2e_Availability!I{src_row}").border = BORDER
    row_out += 1

# Section B rows: b_data_start..row_cursor-1
for src_row in range(b_data_start, row_cursor):
    ws.cell(row=row_out, column=1, value="M2e_Availability!B (string proxy)").border = BORDER
    ws.cell(row=row_out, column=2, value="M2e_string_proxy").border = BORDER
    ws.cell(row=row_out, column=3, value=f"=M2e_Availability!A{src_row}").border = BORDER
    ws.cell(row=row_out, column=4, value=f"=M2e_Availability!B{src_row}").border = BORDER
    ws.cell(row=row_out, column=5, value=f"=M2e_Availability!G{src_row}").border = BORDER
    ws.cell(row=row_out, column=6, value=f"=M2e_Availability!F{src_row}").border = BORDER
    ws.cell(row=row_out, column=6).number_format = "0.00"
    ws.cell(row=row_out, column=7, value="(see Config)").border = BORDER
    ws.cell(row=row_out, column=8, value=f"=M2e_Availability!I{src_row}").border = BORDER
    row_out += 1

# Conditional fill severity
sev_range = f"E5:E{row_out-1}"
for sev_name, fill in SEV_FILL.items():
    ws.conditional_formatting.add(
        sev_range,
        CellIsRule(operator="equal", formula=[f'"{sev_name}"'], fill=fill),
    )

# AutoFilter so user can filter out NORMAL/EMPTY
ws.auto_filter.ref = f"A4:H{row_out-1}"

for letter, w in zip("ABCDEFGH", [28, 18, 14, 10, 13, 18, 12, 70]):
    ws.column_dimensions[letter].width = w
ws.freeze_panes = "A5"


# ===========================================================================
# Save
# ===========================================================================
OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)
print(f"Workbook saved: {OUTPUT}")
print(f"Sheets: {wb.sheetnames}")
